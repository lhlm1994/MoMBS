# -*- coding:utf-8 -*-
import os
import torch
import torch.nn.functional as F
from torch.autograd import Variable
from data.datasets import input_dataset
from models import *
import argparse
from torch.utils.data import Sampler
import math
import openpyxl


parser = argparse.ArgumentParser()
parser.add_argument('--lr', type = float, default = 0.1)
parser.add_argument('--noise_type', type = str, help='clean, aggre, worst, rand1, rand2, rand3, clean100, noisy100', default='clean')
parser.add_argument('--noise_path', type = str, help='path of CIFAR-10_human.pt', default=None)
parser.add_argument('--dataset', type = str, help = ' cifar10 or cifar100', default = 'cifar100')
parser.add_argument('--n_epoch', type=int, default=300)
parser.add_argument('--seed', type=int, default=0)
parser.add_argument('--print_freq', type=int, default=50)
parser.add_argument('--num_workers', type=int, default=4, help='how many subprocesses to use for data loading')
parser.add_argument('--is_human', action='store_true', default=True)


uncertain_mode=1 # 1: under trub , 0 without turb
uncertaint_T=2# 4 8
Pre_EPOCH=20
# uncertainty and loss are unstabel in the early stage
pair_mode=3 # 1 loss or uncertainty l+h,
if pair_mode==1 or pair_mode ==2:
    loss_only=1 # 0: use uncertainty only 1 use loss only
else:
    loss_only = 0
# BATCH_SIZE=128
# BATCH_SIZE=256
BATCH_SIZE=32

work_dir=r'/home1/hli/project_CIFA_N/cifar-10-100n/ul_lh_human_noise_weight_b32_pre20_T2/'

###uncertain and loss pairing

if not os.path.exists(work_dir):
    os.mkdir(work_dir)
def compute_uncertainty(model, T, x,label):
    model.eval()
    with torch.set_grad_enabled(False):  # save memory during inference
        for t in range(uncertaint_T):
            with torch.no_grad():
                # logits1, logits = model(features)
                if uncertain_mode==0:
                    logits1= model(x)
                else:
                    logits1 = model(x,noise=True)
                B = logits1.shape[0]
                gt_logits1=torch.tensor([logits1[i][label[i]] for i in range(B)])
                logits1=gt_logits1
                logits1 -= logits1.min()
                logits1 = torch.nn.functional.normalize(logits1,dim=0)

                # there is still one batch input, noise is added in the forward process,
                if t == 0:
                    output_feature = torch.zeros([T * B]).cuda()
                    # add a new container with shape of T*B , c ,w ,h  in the first time
                    output_feature[B * t: B *  (t + 1)] = logits1
                else:
                    output_feature[B * t: B *  (t + 1)] = logits1
        output_feature_r = output_feature.reshape(T , B)
        output_feature_r = output_feature_r.mean(0)  # get mean of T times
        uncertainty = -1.0 * output_feature_r * torch.log(output_feature_r + 1e-6)

        model.train()
        return uncertainty

def Uncertainty_pair(uncertainty,img2id=None,pair_mode=1,loss_perimg=None):
    # uncertainty_sorted = dict(sorted(uncertainty.items(), key=lambda x: x[1], reverse=False))
    uncertainty_sorted = sorted(uncertainty.items(), key=lambda x: x[1], reverse=False)# use list to get slices, dict is not capable
    loss_sorted = sorted(loss_perimg.items(), key=lambda x: x[1], reverse=False)# use list to get slices, dict is not capable

    len_dict=len(uncertainty_sorted)
    if pair_mode==1:#low+hi
        mid=int(len_dict/2)
        top_dict=uncertainty_sorted[:mid]#find top mid samples
        bottom_dict=uncertainty_sorted[mid:]
        reversed_bottom_dict=sorted(bottom_dict, key=lambda x :x[1],reverse=True)
        if len_dict%2==1:
            top_dict.append(reversed_bottom_dict[-1])
            # if the len is odd(i.e., the first sample in bottom dict), add the mid uncertainty sample into the top dict\
        indices=[]
        for t_img, b_img in zip(top_dict,reversed_bottom_dict):
            t_img_name=t_img[0]
            b_img_name=b_img[0]
            indices.append(t_img_name)
            indices.append(b_img_name)
        return indices
    elif pair_mode==2: #low2hi
        indices = []
        for img in uncertainty_sorted:
            img_name = img[0]
            # indices.append(img2id[img_name])
            indices.append(img_name)
        return indices
    elif pair_mode == 3:  # loss +u low hi
        uncertainty_name_sorted = [u_tuple[0] for u_tuple in uncertainty_sorted]
        loss_name_sorted = [l_tuple[0] for l_tuple in loss_sorted]
        uncertainty_loss=[[name,uncertainty_name_sorted.index(name)+loss_name_sorted.index(name)] for name in loss_name_sorted ]
        uncertainty_loss_sorted= sorted(uncertainty_loss, key=lambda x: x[1], reverse=False)
        mid = int(len(uncertainty_loss_sorted) / 2)
        top_dict = uncertainty_loss_sorted[:mid]  # find top mid samples
        bottom_dict = uncertainty_loss_sorted[mid:]
        reversed_bottom_dict = sorted(bottom_dict, key=lambda x: x[1], reverse=True)
        if len_dict % 2 == 1:
            top_dict.append(reversed_bottom_dict[-1])
            # if the len is odd(i.e., the first sample in bottom dict), add the mid uncertainty sample into the top dict\
        indices = []
        for t_img, b_img in zip(top_dict, reversed_bottom_dict):
            t_img_name = t_img[0]
            b_img_name = b_img[0]
            # indices.append(img2id[t_img_name])
            # indices.append(img2id[b_img_name])
            indices.append(t_img_name)
            indices.append(b_img_name)
        return indices
    elif pair_mode == 4:  # loss +u low2hi
        uncertainty_name_sorted = [u_tuple[0] for u_tuple in uncertainty_sorted]
        loss_name_sorted = [l_tuple[0] for l_tuple in loss_sorted]
        uncertainty_loss = [[name, uncertainty_name_sorted.index(name) + loss_name_sorted.index(name)] for name in
                            loss_name_sorted]
        uncertainty_loss_sorted = sorted(uncertainty_loss, key=lambda x: x[1], reverse=False)
        indices = []
        for img in uncertainty_loss_sorted:
            img_name = img[0]

            indices.append(img_name)
        return indices

def uncertain2index(uncertaity,img_list=None,loss_perimg=None):
    # img2id={img:i for i, img in enumerate(img_list)}
    # indices=Uncertainty_pair(uncertainty=uncertaity,img2id=img2id,pair_mode=pair_mode,loss_perimg=loss_perimg)
    indices=Uncertainty_pair(uncertainty=uncertaity,pair_mode=pair_mode,loss_perimg=loss_perimg)
    #pair mode 1: u low+hi, 2: u low low hihi 3. loss+u low hi
    return indices
class Uncertainty_GroupSampler(Sampler):
    """Sampler that restricts data loading to a subset of the dataset.
        It is especially useful in conjunction with
        :class:`torch.nn.parallel.DistributedDataParallel`. In such case, each
        process can pass a DistributedSampler instance as a DataLoader sampler,
        and load a subset of the original dataset that is exclusive to it.
        .. note::
            Dataset is assumed to be of constant size.
        Arguments:
            dataset: Dataset used for sampling.
            num_replicas (optional): Number of processes participating in
                distributed training.
            rank (optional): Rank of the current process within num_replicas.
        """

    def __init__(self,
                 dataset,
                 epoch,
                 samples_per_gpu=1,
                 num_replicas=None,
                 rank=None,
                 uncertainty=None,
                 img_list=None,
                 cls_losses_perimg=None):
        if num_replicas is None:
            num_replicas = 1
        if rank is None:
            rank = 0
        self.dataset = dataset
        self.samples_per_gpu = samples_per_gpu
        self.num_replicas = num_replicas# self.num_replicas 是distrubted GPU的数量
        self.rank = rank
        self.epoch = epoch
        self.uncertainty=uncertainty
        self.cls_losses_perimg=cls_losses_perimg
        self.img_list=img_list

        self.group_sizes = len(self.dataset)# the num of data

        self.num_samples = 0

        self.total_size = int(math.ceil(self.group_sizes / self.samples_per_gpu))

    def __iter__(self):
        g = torch.Generator()
        g.manual_seed(self.epoch)
        indice = uncertain2index(self.uncertainty,loss_perimg=self.cls_losses_perimg)  # han add
        if (self.epoch-1)%10==0:
            file_path= os.path.join(work_dir,'loss_uncertainty_perimg.xlsx')
            if not os.path.exists(file_path):
                wb=openpyxl.Workbook()
            else:
                wb =openpyxl.load_workbook(file_path)
            num_activeSheet =len(wb.sheetnames)
            mySheet = wb.create_sheet(index=num_activeSheet, title=str(num_activeSheet)+'_'+str(self.epoch))
            row=1
            for k in self.uncertainty.keys():
                # uncertainty=self.uncertainty[k].item()
                uncertainty=self.uncertainty[k]
                loss=self.cls_losses_perimg[k]
                mySheet.cell(row=row,column=1).value=k
                mySheet.cell(row=row,column=2).value=uncertainty
                mySheet.cell(row=row,column=3).value=loss
                row+=1
            wb.save(file_path)

        indices = [ # Random the samples_per_gpu img pair, the sort within each img pair remain unchanged
            indice[j] for i in list(
                torch.randperm(
                    len(indice) // self.samples_per_gpu, generator=g))
            for j in range(i * self.samples_per_gpu, (i + 1) *
                           self.samples_per_gpu)
        ]

        #no subset or rank here,
        print(len(indices))
        # print(len(self.uncertainty))
        # print(indices)
        return iter(indices)

    def __len__(self):
        return self.num_samples

    def set_epoch(self, epoch):
        self.epoch = epoch


# Adjust learning rate and for SGD Optimizer

def adjust_learning_rate(optimizer, epoch,alpha_plan):
    for param_group in optimizer.param_groups:
        param_group['lr']=alpha_plan[epoch]
        

def accuracy(logit, target, topk=(1,)):
    """Computes the precision@k for the specified values of k"""
    output = F.softmax(logit, dim=1)
    maxk = max(topk)
    batch_size = target.size(0)

    _, pred = output.topk(maxk, 1, True, True)# _ is the value list pred is the id list
    pred = pred.t()
    correct = pred.eq(target.reshape(1, -1).expand_as(pred))

    res = []
    for k in topk:
        correct_k = correct[:k].reshape(-1).float().sum(0, keepdim=True)
        res.append(correct_k.mul_(100.0 / batch_size))
    return res

# Train the Model
def train(epoch, train_loader, model, optimizer, loss_dic, uncertainty_dic):
    train_total=0
    train_correct=0

    for i, (images, labels, indexes) in enumerate(train_loader):
        ind=indexes.cpu().numpy().transpose()
        batch_size = len(ind)
       
        images = Variable(images).cuda()
        labels = Variable(labels).cuda()
       
        # Forward + Backward + Optimize
        logits = model(images)

        prec, _ = accuracy(logits, labels, topk=(1, 5))# prec is the top1 acc,
        # prec = 0.0
        train_total+=1
        train_correct+=prec
        # loss = F.cross_entropy(logits, labels, reduce = True) # reduce will average among batch
        loss_tem = F.cross_entropy(logits, labels, reduce = False)
        loss=loss_tem.mean(0)
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        if (i+1) % args.print_freq == 0:
            print ('Epoch [%d/%d], Iter [%d/%d] Training Accuracy: %.4F, Loss: %.4f'
                  %(epoch+1, args.n_epoch, i+1, len(train_dataset)//batch_size, prec, loss.data))
        #upadate loss and uncertainty
        tem_uncertainty=compute_uncertainty(model,uncertaint_T,images,labels)
        tem_uncertainty_dict={}
        for i in range(len(loss_tem)):
            loss_dic[int(indexes[i].cpu().detach().numpy())] = float(loss_tem[i].cpu().detach().numpy())
            tem_uncertainty_dict[int(indexes[i].cpu().detach().numpy())] = float(tem_uncertainty[i].cpu().detach().numpy())
        uncertainty_dic.update(tem_uncertainty_dict)
    if loss_only:
        uncertainty_dic = {}  # if use loss only
        uncertainty_dic.update(loss_dic)
    train_acc=float(train_correct)/float(train_total)
    return train_acc

# Evaluate the Model
def evaluate(eval_loader, model,best_acc_,best_epoch):
    model.eval()    # Change model to 'eval' mode.
    print('eval_previous_best epoch%d, best acc %f'% (best_epoch,best_acc_))
    correct = 0
    total = 0
    for images, labels, _ in eval_loader:
        images = Variable(images).cuda()
        logits = model(images)
        outputs = F.softmax(logits, dim=1)
        _, pred = torch.max(outputs.data, 1)
        total += labels.size(0)
        correct += (pred.cpu() == labels).sum()
    acc = 100*float(correct)/float(total)

    return acc



#####################################main code ################################################
args = parser.parse_args()
# Seed
torch.manual_seed(args.seed)
torch.cuda.manual_seed(args.seed)

# Hyper Parameters
# batch_size = 128
learning_rate = args.lr
noise_type_map = {'clean':'clean_label', 'worst': 'worse_label', 'aggre': 'aggre_label', 'rand1': 'random_label1', 'rand2': 'random_label2', 'rand3': 'random_label3', 'clean100': 'clean_label', 'noisy100': 'noisy_label'}
args.noise_type = noise_type_map[args.noise_type]
# load dataset
if args.noise_path is None:
    if args.dataset == 'cifar10':
        args.noise_path = './data/CIFAR-10_human.pt'
    elif args.dataset == 'cifar100':
        args.noise_path = './data/CIFAR-100_human.pt'
    else: 
        raise NameError(f'Undefined dataset {args.dataset}')


train_dataset,eval_dataset,num_classes,num_training_samples = input_dataset(args.dataset,args.noise_type, args.noise_path, args.is_human)
noise_prior = train_dataset.noise_prior
noise_or_not = train_dataset.noise_or_not
print('train_labels:', len(train_dataset.train_labels), train_dataset.train_labels[:10])
# load model
print('building model...')
model = ResNet34(num_classes)
print('building model done')
optimizer = torch.optim.SGD(model.parameters(), lr=learning_rate, weight_decay=0.0005, momentum=0.9)

train_loader = torch.utils.data.DataLoader(dataset = train_dataset,
                                   batch_size = BATCH_SIZE,
                                   num_workers=args.num_workers,
                                   shuffle=True)




eval_loader = torch.utils.data.DataLoader(dataset=eval_dataset,
                                  batch_size = 64,
                                  num_workers=args.num_workers,
                                  shuffle=False)
alpha_plan = [0.1] * 60 + [0.01] * 300
model.cuda()


epoch=0
train_acc = 0

# training
noise_prior_cur = noise_prior
best_acc=0
best_epoch=-1
loss_dic = {}
uncertainty_dic = {}
for epoch in range(args.n_epoch):
# train models
    print(f'epoch {epoch}')
    adjust_learning_rate(optimizer, epoch, alpha_plan)
    if epoch > Pre_EPOCH:

        sampler = Uncertainty_GroupSampler(train_dataset, epoch, samples_per_gpu=BATCH_SIZE, uncertainty=uncertainty_dic,
                                           cls_losses_perimg=loss_dic)
        train_loader = torch.utils.data.DataLoader(train_dataset, batch_size=BATCH_SIZE, sampler=sampler)
    loss_dic = {}
    uncertainty_dic = {}
    model.train()
    train_acc = train(epoch, train_loader, model, optimizer, loss_dic, uncertainty_dic)
    # evaluate models

    eval_acc = evaluate(eval_loader=eval_loader, model=model,best_acc_=best_acc,best_epoch=best_epoch)
    if eval_acc >best_acc:
        best_acc=eval_acc
        best_epoch=epoch
    # save results
    print('train acc on train images is ', train_acc)
    print('eval acc on eval images is ', eval_acc)
